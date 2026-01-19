import os
import sys
import datetime
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import telebot
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from bot_state import init_state, is_active as check_bot_active

# Import singleton manager untuk mencegah bot duplikat
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from bot_singleton import BotSingleton
    SINGLETON_AVAILABLE = True
except ImportError:
    SINGLETON_AVAILABLE = False
    print("⚠️  bot_singleton.py tidak ditemukan, singleton protection disabled")

load_dotenv()
TOKEN = os.getenv('TELEGRAM_TOKEN')
if not TOKEN:
    raise SystemExit('Missing TELEGRAM_TOKEN in environment. Create a .env file or set the variable.')
TARGET_CHAT_ID = os.getenv('TARGET_CHAT_ID')  # Group chat ID to receive daily report
try:
    MAX_BREAK_USERS = int(os.getenv('MAX_BREAK_USERS', '4'))
except Exception:
    MAX_BREAK_USERS = 4

bot = telebot.TeleBot(TOKEN)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, 'attendance.xlsx')
SESSIONS_PATH = os.path.join(BASE_DIR, 'sessions.json')
ALLOWED_UPDATES = ['message', 'callback_query']  # ensure we also fetch pending updates


def safe_log(message: str):
    """Console logging that won't crash on non-ASCII (Windows/GBK)."""
    try:
        print(message)
    except Exception:
        try:
            safe_msg = message.encode('ascii', 'backslashreplace').decode('ascii')
            print(safe_msg)
        except Exception:
            # last resort: drop logging
            pass


def safe_reply(msg, text: str, retries: int = 2, delay: float = 1.0, **kwargs):
    """Reply with retry and swallow network errors."""
    for _ in range(retries + 1):
        try:
            return bot.reply_to(msg, text, **kwargs)
        except Exception as e:
            safe_log(f"Reply failed: {e}")
            try:
                import time
                time.sleep(delay)
            except Exception:
                pass
    return None


def set_bot_commands():
    """Register all bot commands for Telegram client menus."""
    bot.set_my_commands([
        telebot.types.BotCommand('start', 'Tampilkan bantuan'),
        telebot.types.BotCommand('help', 'Tampilkan bantuan'),
        telebot.types.BotCommand('masuk', 'Catat masuk kerja'),
        telebot.types.BotCommand('pulang', 'Catat pulang & hitung total'),
        telebot.types.BotCommand('wc', 'Mulai istirahat WC'),
        telebot.types.BotCommand('merokok', 'Mulai istirahat merokok'),
        telebot.types.BotCommand('kembali_kursi', 'Akhiri istirahat'),
        telebot.types.BotCommand('info', 'Lihat info absensi hari ini'),
        telebot.types.BotCommand('status', 'Cek status bot'),
        telebot.types.BotCommand('capacity', 'Cek kapasitas istirahat'),
        telebot.types.BotCommand('reset_activity', 'Nolkan hitung istirahat saya'),
        telebot.types.BotCommand('reset', 'Reset session saya hari ini'),
        telebot.types.BotCommand('clear_capacity', 'Kosongkan kapasitas (admin)'),
        telebot.types.BotCommand('reset_all_sessions', 'Hapus seluruh sesi (admin)'),
        telebot.types.BotCommand('off', 'Matikan bot'),
        telebot.types.BotCommand('on', 'Hidupkan bot'),
    ])

# Per-user session state in memory
sessions = {}
# Global concurrency limits for active breaks (max 2 users each)
active_wc_users = set()
active_merokok_users = set()
# Initialize bot state file
init_state()

def load_sessions():
    try:
        import json
        if os.path.exists(SESSIONS_PATH):
            with open(SESSIONS_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # restore sessions dict
            global sessions, active_wc_users, active_merokok_users
            sessions = {}
            for uid_str, s in data.get('sessions', {}).items():
                uid = int(uid_str)
                # restore datetimes from iso strings
                masuk = datetime.datetime.fromisoformat(s['masuk']) if s.get('masuk') else None
                pulang = datetime.datetime.fromisoformat(s['pulang']) if s.get('pulang') else None
                break_start = datetime.datetime.fromisoformat(s['break_start']) if s.get('break_start') else None
                sessions[uid] = {
                    'masuk': masuk,
                    'pulang': pulang,
                    'break_active': s.get('break_active'),
                    'break_start': break_start,
                    'merokok_count': int(s.get('merokok_count', 0)),
                    'merokok_dur': datetime.timedelta(seconds=int(s.get('merokok_dur_sec', 0))),
                    'wc_count': int(s.get('wc_count', 0)),
                    'wc_dur': datetime.timedelta(seconds=int(s.get('wc_dur_sec', 0))),
                    'user_name': s.get('user_name', 'Unknown'),
                }
            active_wc_users = set(map(int, data.get('active_wc_users', [])))
            active_merokok_users = set(map(int, data.get('active_merokok_users', [])))
        cleanup_active_sets()
    except Exception:
        pass

def save_sessions():
    try:
        import json
        out = {
            'sessions': {},
            'active_wc_users': list(active_wc_users),
            'active_merokok_users': list(active_merokok_users),
        }
        for uid, s in sessions.items():
            out['sessions'][str(uid)] = {
                'masuk': s['masuk'].isoformat() if s.get('masuk') else None,
                'pulang': s['pulang'].isoformat() if s.get('pulang') else None,
                'break_active': s.get('break_active'),
                'break_start': s['break_start'].isoformat() if s.get('break_start') else None,
                'merokok_count': s.get('merokok_count', 0),
                'merokok_dur_sec': int(s.get('merokok_dur', datetime.timedelta()).total_seconds()),
                'wc_count': s.get('wc_count', 0),
                'wc_dur_sec': int(s.get('wc_dur', datetime.timedelta()).total_seconds()),
                'user_name': s.get('user_name', 'Unknown'),
            }
        with open(SESSIONS_PATH, 'w', encoding='utf-8') as f:
            json.dump(out, f)
    except Exception:
        pass


def cleanup_active_sets():
    """Remove stale active break users that no longer have an active session or are stuck too long."""
    now = now_jkt()
    stale_wc = set()
    stale_mr = set()

    for uid in list(active_wc_users):
        sess = sessions.get(uid)
        if not sess or sess.get('break_active') != 'wc':
            stale_wc.add(uid)
            continue
        # Auto-close WC break if it has been open past midnight or >6 hours
        start = sess.get('break_start')
        if start and (start.date() != now.date() or (now - start) > datetime.timedelta(hours=6)):
            sess['wc_dur'] += now - start
            sess['break_active'] = None
            sess['break_start'] = None
            stale_wc.add(uid)

    for uid in list(active_merokok_users):
        sess = sessions.get(uid)
        if not sess or sess.get('break_active') != 'merokok':
            stale_mr.add(uid)
            continue
        # Auto-close merokok break if it has been open past midnight or >6 hours
        start = sess.get('break_start')
        if start and (start.date() != now.date() or (now - start) > datetime.timedelta(hours=6)):
            sess['merokok_dur'] += now - start
            sess['break_active'] = None
            sess['break_start'] = None
            stale_mr.add(uid)

    if stale_wc:
        active_wc_users.difference_update(stale_wc)
    if stale_mr:
        active_merokok_users.difference_update(stale_mr)
    if stale_wc or stale_mr:
        save_sessions()
REST_QUOTA = datetime.timedelta(hours=1)  # total allowed rest per workday
try:
    JKT = ZoneInfo("Asia/Jakarta")
except Exception:
    JKT = None

def now_jkt() -> datetime.datetime:
    if JKT is not None:
        return datetime.datetime.now(tz=JKT)
    # Fallback: naive local time if tzdata not available yet
    return datetime.datetime.now()


def fmt_time(dt: datetime.datetime) -> str:
    try:
        if JKT is not None:
            local = dt.astimezone(JKT) if dt.tzinfo else dt
        else:
            local = dt
        return local.strftime('%H:%M:%S')
    except Exception:
        return dt.strftime('%H:%M:%S')

def fmt_time_wib(dt: datetime.datetime) -> str:
    return f"{fmt_time(dt)} WIB"


def fmt_delta(td: datetime.timedelta) -> str:
    total_seconds = int(td.total_seconds())
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def ensure_workbook():
    if not os.path.exists(XLSX_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        ws.append([
            'Tanggal', 'User ID', 'Nama', 'Masuk', 'Pulang',
            'Merokok (kali)', 'Merokok (durasi)',
            'WC (kali)', 'WC (durasi)',
            'Waktu Kerja', 'Total Istirahat', 'Sisa Istirahat'
        ])
        # Optional: set column widths
        for i in range(1, 13):
            ws.column_dimensions[get_column_letter(i)].width = 18
        wb.save(XLSX_PATH)


def append_summary_to_xlsx(user_id: int, name: str, masuk: datetime.datetime, pulang: datetime.datetime,
                           merokok_count: int, merokok_dur: datetime.timedelta,
                           wc_count: int, wc_dur: datetime.timedelta,
                           kerja: datetime.timedelta, rest_total: datetime.timedelta,
                           rest_sisa: datetime.timedelta):
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    ws.append([
        now_jkt().strftime('%Y-%m-%d'),
        user_id,
        name,
        fmt_time(masuk) if masuk else '',
        fmt_time(pulang) if pulang else '',
        merokok_count,
        fmt_delta(merokok_dur),
        wc_count,
        fmt_delta(wc_dur),
        fmt_delta(kerja),
        fmt_delta(rest_total),
        fmt_delta(rest_sisa),
    ])
    wb.save(XLSX_PATH)


def get_user_name(user) -> str:
    return user.first_name or user.username or 'User'


@bot.message_handler(commands=['start', 'help'])
def send_help(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    # Set bot command menu in Telegram clients
    set_bot_commands()

    # Send a reply keyboard for quick taps
    kb = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.row(
        telebot.types.KeyboardButton('/masuk'),
        telebot.types.KeyboardButton('/pulang')
    )
    kb.row(
        telebot.types.KeyboardButton('/wc'),
        telebot.types.KeyboardButton('/merokok'),
        telebot.types.KeyboardButton('/kembali_kursi')
    )
    kb.row(
        telebot.types.KeyboardButton('/off'),
        telebot.types.KeyboardButton('/on')
    )

    text = (
        'Bot Absensi — perintah yang tersedia:\n'
        '/masuk — catat masuk kerja\n'
        '/pulang — catat pulang (hitung total jam kerja, reset)\n'
        '/wc — mulai istirahat WC\n'
        '/merokok — mulai istirahat merokok\n'
        '/kembali_kursi — akhiri istirahat (wc/merokok)\n'
        '/info — lihat informasi absensi hari ini\n'
        '/reset — reset session hari ini (konfirmasi diperlukan)\n'
        '/status — cek status bot\n'
        '/capacity — cek kapasitas istirahat\n'
        '/off — matikan bot\n'
        '/on — hidupkan bot\n'
    )
    safe_reply(message, text, reply_markup=kb)


@bot.message_handler(commands=['status'])
def status(message):
    status_text = "✅ Bot Aktif (ON)" if check_bot_active() else "⛔ Bot Mati (OFF)"
    safe_reply(message, f"Build: session+xlsx v1.0 — siap.\n{status_text}")


@bot.message_handler(commands=['info'])
def info(message):
    """Menampilkan informasi bot absensi"""
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    
    user = message.from_user
    user_session = sessions.get(user.id, {})
    
    masuk = user_session.get('masuk')
    pulang = user_session.get('pulang')
    break_active = user_session.get('break_active')
    merokok_count = user_session.get('merokok_count', 0)
    merokok_dur = user_session.get('merokok_dur', datetime.timedelta(0))
    wc_count = user_session.get('wc_count', 0)
    wc_dur = user_session.get('wc_dur', datetime.timedelta(0))
    
    # Build info text
    info_text = f"📊 Informasi Absensi - {get_user_name(user)}\n\n"
    info_text += f"Status Hari Ini:\n"
    
    if masuk:
        info_text += f"✅ Masuk: {fmt_time_wib(masuk)}\n"
    else:
        info_text += f"⏳ Belum masuk\n"
    
    if pulang:
        info_text += f"✅ Pulang: {fmt_time_wib(pulang)}\n"
    else:
        info_text += f"⏳ Belum pulang\n"
    
    info_text += f"\n📈 Aktivitas Istirahat:\n"
    info_text += f"🚽 WC: {wc_count}x ({fmt_delta(wc_dur)})\n"
    info_text += f"🚬 Merokok: {merokok_count}x ({fmt_delta(merokok_dur)})\n"
    
    if break_active:
        info_text += f"\n🔄 Istirahat Aktif: {break_active.upper()}\n"
    
    info_text += f"\n⚙️ Bot Status: {'🟢 ON' if check_bot_active() else '⚫ OFF'}\n"
    
    safe_reply(message, info_text)


@bot.message_handler(commands=['reset'])
def reset_session(message):
    """Reset session user (hanya untuk admin/ownership)"""
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    
    user = message.from_user
    
    # Check if user has a session
    if user.id not in sessions:
        safe_reply(message, "❌ Anda belum memiliki session hari ini.")
        return
    
    # Confirm reset
    kb = telebot.types.InlineKeyboardMarkup()
    kb.add(
        telebot.types.InlineKeyboardButton("✅ Ya, reset", callback_data=f"reset_confirm_{user.id}"),
        telebot.types.InlineKeyboardButton("❌ Batal", callback_data=f"reset_cancel_{user.id}")
    )
    
    bot.send_message(
        message.chat.id,
        "⚠️ Reset akan menghapus semua data absensi hari ini untuk user Anda.\n"
        "Apakah Anda yakin?",
        reply_markup=kb
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith('reset_confirm_'))
def confirm_reset(call):
    """Konfirmasi reset"""
    try:
        user_id = int(call.data.split('_')[2])
        
        if user_id == call.from_user.id:
            if user_id in sessions:
                sessions[user_id] = {
                    'masuk': None,
                    'pulang': None,
                    'break_active': None,
                    'break_start': None,
                    'merokok_count': 0,
                    'merokok_dur': datetime.timedelta(0),
                    'wc_count': 0,
                    'wc_dur': datetime.timedelta(0),
                    'user_name': sessions[user_id].get('user_name', f'User {user_id}'),
                }
                # Remove from active sets and persist
                active_wc_users.discard(user_id)
                active_merokok_users.discard(user_id)
                save_sessions()
                
                bot.answer_callback_query(call.id, "✅ Session reset berhasil!", show_alert=False)
                bot.send_message(call.message.chat.id, "✅ Data absensi Anda hari ini sudah direset.\nSilakan mulai dengan /masuk untuk masuk kerja.")
            else:
                bot.answer_callback_query(call.id, "❌ Session tidak ditemukan", show_alert=False)
        else:
            bot.answer_callback_query(call.id, "❌ Anda tidak berhak reset session orang lain", show_alert=True)
    except Exception as e:
        bot.answer_callback_query(call.id, f"❌ Error: {str(e)}", show_alert=True)


@bot.callback_query_handler(func=lambda call: call.data.startswith('reset_cancel_'))
def cancel_reset(call):
    """Cancel reset"""
    try:
        user_id = int(call.data.split('_')[2])
        if user_id == call.from_user.id:
            bot.answer_callback_query(call.id, "Reset dibatalkan", show_alert=False)
            bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception:
        pass


@bot.message_handler(commands=['capacity'])
def capacity(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    cleanup_active_sets()
    
    wc_used = len(active_wc_users)
    mr_used = len(active_merokok_users)
    
    reply = f"📊 Kapasitas Istirahat Saat Ini:\n\n"
    
    # WC Status
    reply += f"🚽 WC: {wc_used}/{MAX_BREAK_USERS} aktif\n"
    if active_wc_users:
        for uid in active_wc_users:
            if uid in sessions:
                user_name = sessions[uid].get('user_name', f'User {uid}')
                reply += f"  👤 {user_name}\n"
    else:
        reply += "  (kosong)\n"
    
    reply += "\n"
    
    # Merokok Status
    reply += f"🚬 Merokok: {mr_used}/{MAX_BREAK_USERS} aktif\n"
    if active_merokok_users:
        for uid in active_merokok_users:
            if uid in sessions:
                user_name = sessions[uid].get('user_name', f'User {uid}')
                reply += f"  👤 {user_name}\n"
    else:
        reply += "  (kosong)\n"
    
    reply += "\n"
    if wc_used >= MAX_BREAK_USERS or mr_used >= MAX_BREAK_USERS:
        reply += "⚠️ Kapasitas penuh! Tunggu hingga ada yang /kembali_kursi."
    else:
        reply += "✅ Masih ada slot istirahat tersedia."
    
    safe_reply(message, reply)


@bot.message_handler(commands=['clear_capacity'])
def clear_capacity(message):
    """Kosongkan daftar kapasitas istirahat jika tersisa nama lama."""
    cleanup_active_sets()
    # Force-clear any remaining active users
    active_wc_users.clear()
    active_merokok_users.clear()
    save_sessions()
    safe_reply(message, "✅ Kapasitas istirahat sudah dikosongkan.")


@bot.message_handler(commands=['reset_all_sessions'])
def reset_all_sessions(message):
    """Hapus seluruh sesi & kapasitas (gunakan hati-hati)."""
    # Simple guard: only allow if sender is TARGET_CHAT_ID owner? For now allow but warn.
    sessions.clear()
    active_wc_users.clear()
    active_merokok_users.clear()
    save_sessions()
    safe_reply(message, "✅ Semua sesi dan kapasitas sudah dihapus. Minta semua user kirim /masuk lagi.")


@bot.message_handler(commands=['reset_activity'])
def reset_activity(message):
    """Reset hitung wc/merokok dan akhiri istirahat aktif milik pengirim."""
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    user = message.from_user
    sess = sessions.get(user.id)
    if not sess or not sess.get('masuk'):
        safe_reply(message, "Silakan kirim /masuk terlebih dahulu.")
        return
    # Akhiri istirahat aktif jika ada
    active_wc_users.discard(user.id)
    active_merokok_users.discard(user.id)
    sess['break_active'] = None
    sess['break_start'] = None
    sess['merokok_count'] = 0
    sess['merokok_dur'] = datetime.timedelta(0)
    sess['wc_count'] = 0
    sess['wc_dur'] = datetime.timedelta(0)
    save_sessions()
    safe_reply(message, "✅ Aktivitas istirahat kamu sudah direset. Mulai lagi dengan /wc atau /merokok bila perlu.")


@bot.message_handler(commands=['masuk'])
def handle_masuk(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    user = message.from_user
    now = now_jkt()
    sessions[user.id] = {
        'masuk': now,
        'pulang': None,
        'break_active': None,  # 'wc' or 'merokok'
        'break_start': None,
        'merokok_count': 0,
        'merokok_dur': datetime.timedelta(0),
        'wc_count': 0,
        'wc_dur': datetime.timedelta(0),
        'user_name': get_user_name(user),  # Store user name for capacity display
    }
    save_sessions()
    reply = (
        f"command masuk dikirim\n"
        f"{get_user_name(user)}\n"
        f"masuk kerja pada pukul ({fmt_time_wib(now)})"
    )
    safe_reply(message, reply)


def rest_remaining(user_id: int) -> datetime.timedelta:
    sess = sessions.get(user_id)
    if not sess:
        return REST_QUOTA
    total = sess['merokok_dur'] + sess['wc_dur']
    return max(datetime.timedelta(0), REST_QUOTA - total)


def start_break(user, kind: str):
    sess = sessions.get(user.id)
    now = now_jkt()
    if not sess or not sess.get('masuk'):
        return "Silakan kirim /masuk terlebih dahulu."
    if sess['break_active']:
        return "Sedang istirahat, kirim /kembali_kursi untuk mengakhiri."
    # Check if user has remaining time
    remaining = rest_remaining(user.id)
    if remaining.total_seconds() <= 0:
        return f"⚠️ Waktu istirahat Anda sudah habis. Sisa waktu: {fmt_delta(remaining)}\nAnda tidak dapat melakukan istirahat lagi hari ini."
    # Enforce max 2 concurrent users per break type (global)
    if kind == 'wc':
        slots_used = len(active_wc_users)
        if user.id not in active_wc_users and slots_used >= MAX_BREAK_USERS:
            return f"WC sedang penuh (maks {MAX_BREAK_USERS} orang). Saat ini {slots_used} orang, tunggu hingga ada yang /kembali_kursi."
        active_wc_users.add(user.id)
        slots_used = len(active_wc_users)
    else:
        slots_used = len(active_merokok_users)
        if user.id not in active_merokok_users and slots_used >= MAX_BREAK_USERS:
            return f"Area merokok sedang penuh (maks {MAX_BREAK_USERS} orang). Saat ini {slots_used} orang, tunggu hingga ada yang /kembali_kursi."
        active_merokok_users.add(user.id)
        slots_used = len(active_merokok_users)
    sess['break_active'] = kind
    sess['break_start'] = now
    if kind == 'wc':
        sess['wc_count'] += 1
    else:
        sess['merokok_count'] += 1
    save_sessions()
    remaining = rest_remaining(user.id)
    return (
        f"command {kind} dikirim\n"
        f"{get_user_name(user)}\n"
        f"pergi {kind} ke {sess['wc_count'] if kind=='wc' else sess['merokok_count']} kali pada pukul ({fmt_time_wib(now)})\n"
        f"Sedang {kind}: {slots_used}/{MAX_BREAK_USERS} orang\n"
        f"batas maximum {MAX_BREAK_USERS} orang untuk melakukan {kind}\n"
        f"sisa waktu ({fmt_delta(remaining)})"
    )


@bot.message_handler(commands=['wc'])
def handle_wc(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    reply = start_break(message.from_user, 'wc')
    safe_reply(message, reply)


@bot.message_handler(commands=['merokok'])
def handle_merokok(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    reply = start_break(message.from_user, 'merokok')
    safe_reply(message, reply)


@bot.message_handler(commands=['kembali_kursi'])
def handle_kembali(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    try:
        user = message.from_user
        sess = sessions.get(user.id)
        now = now_jkt()
        
        # Validation checks
        if not sess or not sess.get('masuk'):
            safe_reply(message, "Silakan kirim /masuk terlebih dahulu.")
            return
        
        if not sess.get('break_active'):
            safe_reply(message, "Tidak sedang istirahat.")
            return
        
        if not sess.get('break_start'):
            safe_reply(message, "⚠️ Error: Data break_start tidak ditemukan. Silakan hubungi admin.")
            safe_log(f"Error: User {user.id} ({get_user_name(user)}) has break_active but no break_start")
            # Clean up the broken state
            sess['break_active'] = None
            if user.id in active_wc_users:
                active_wc_users.discard(user.id)
            if user.id in active_merokok_users:
                active_merokok_users.discard(user.id)
            save_sessions()
            return
        
        kind = sess['break_active']
        start = sess['break_start']
        dur = now - start
        
        # Update duration based on break type
        if kind == 'wc':
            sess['wc_dur'] += dur
            count = sess['wc_count']
            total_used = sess['wc_dur']
        else:
            sess['merokok_dur'] += dur
            count = sess['merokok_count']
            total_used = sess['merokok_dur']
        
        # End break
        sess['break_active'] = None
        sess['break_start'] = None
        
        # Remove from active sets
        if kind == 'wc':
            active_wc_users.discard(user.id)
        else:
            active_merokok_users.discard(user.id)
        
        save_sessions()
        remaining = rest_remaining(user.id)
        
        reply = (
            f"command kembali_kursi dikirim\n"
            f"{get_user_name(user)}\n"
            f"kembali dari {kind} total {count} kali dengan waktu yang digunakan ({fmt_delta(total_used)})\n"
            f"sisa waktu ({fmt_delta(remaining)})"
        )
        safe_reply(message, reply)
        # Log in ASCII to avoid encoding issues on some Windows consoles
        safe_log(f"OK User {user.id} ({get_user_name(user)}) returned from {kind}, remaining: {fmt_delta(remaining)}")
        
    except Exception as e:
        error_msg = "⚠️ Terjadi kesalahan saat memproses kembali_kursi. Silakan coba lagi atau hubungi admin."
        safe_reply(message, error_msg)
        safe_log(f"ERROR in handle_kembali for user {message.from_user.id}: {str(e)}")
        import traceback
        traceback.print_exc()


@bot.message_handler(commands=['pulang'])
def handle_pulang(message):
    if not check_bot_active():
        safe_reply(message, "Bot sedang dimatikan. Gunakan /on untuk mengaktifkan kembali.")
        return
    user = message.from_user
    sess = sessions.get(user.id)
    now = now_jkt()
    if not sess or not sess.get('masuk'):
        safe_reply(message, "Silakan kirim /masuk terlebih dahulu.")
        return
    # if on break, close it first
    if sess['break_active'] and sess['break_start']:
        dur = now - sess['break_start']
        if sess['break_active'] == 'wc':
            sess['wc_dur'] += dur
            active_wc_users.discard(user.id)
        else:
            sess['merokok_dur'] += dur
            active_merokok_users.discard(user.id)
        sess['break_active'] = None
        sess['break_start'] = None

    masuk = sess['masuk']
    pulang = now
    total_rest = sess['wc_dur'] + sess['merokok_dur']
    kerja = (pulang - masuk) - total_rest
    rest_sisa = max(datetime.timedelta(0), REST_QUOTA - total_rest)

    reply = (
        f"command pulang dikirim\n"
        f"{get_user_name(user)}\n"
        f"masuk ({fmt_time_wib(masuk)})\n"
        f"pulang ({fmt_time_wib(pulang)})\n"
        f"merokok {sess['merokok_count']} kali ({fmt_delta(sess['merokok_dur'])})\n"
        f"wc {sess['wc_count']} kali ({fmt_delta(sess['wc_dur'])})\n"
        f"waktu kerja ({fmt_delta(kerja)})\n"
        f"total waktu istirahat ({fmt_delta(total_rest)})\n"
        f"sisa waktu istirahat ({fmt_delta(rest_sisa)})"
    )
    safe_reply(message, reply)

    # Append to Excel
    append_summary_to_xlsx(
        user_id=user.id,
        name=get_user_name(user),
        masuk=masuk,
        pulang=pulang,
        merokok_count=sess['merokok_count'],
        merokok_dur=sess['merokok_dur'],
        wc_count=sess['wc_count'],
        wc_dur=sess['wc_dur'],
        kerja=kerja,
        rest_total=total_rest,
        rest_sisa=rest_sisa,
    )

    # Reset session
    sessions.pop(user.id, None)
    save_sessions()


@bot.message_handler(commands=['off'])
def handle_off(message):
    from bot_state import set_active
    set_active(False)
    safe_reply(message, "✅ Bot telah dimatikan (OFF). Bot masih berjalan tapi tidak akan merespons perintah apapun.\nGunakan /on untuk menghidupkan kembali.")
    safe_log(f'Bot dimatikan oleh {get_user_name(message.from_user)} pada {fmt_time_wib(now_jkt())}')


@bot.message_handler(commands=['on'])
def handle_on(message):
    from bot_state import set_active
    set_active(True)
    safe_reply(message, "✅ Bot telah diaktifkan (ON). Bot siap menerima perintah.")
    safe_log(f'Bot diaktifkan oleh {get_user_name(message.from_user)} pada {fmt_time_wib(now_jkt())}')


if __name__ == '__main__':
    # Singleton check - mencegah bot duplikat
    singleton = None
    if SINGLETON_AVAILABLE:
        singleton = BotSingleton('absen')
        if not singleton.acquire():
            print("❌ Bot Absen sudah berjalan! Keluar...")
            sys.exit(1)

    try:
        ensure_workbook()
        load_sessions()
        set_bot_commands()
        safe_log('Bot Absensi berjalan — menunggu pesan...')

        # Start scheduler thread to send excel at 21:00 WIB daily
        import threading
        import time

        def send_daily_report_loop():
            last_sent_date = None
            while True:
                try:
                    now = now_jkt()
                    # Send once per day at 21:00 WIB
                    if now.hour == 21 and now.minute == 0:
                        today_str = now.strftime('%Y-%m-%d')
                        if last_sent_date != today_str:
                            if TARGET_CHAT_ID:
                                try:
                                    with open(XLSX_PATH, 'rb') as f:
                                        bot.send_document(int(TARGET_CHAT_ID), f, caption=f'Rekap absensi {today_str}')
                                    last_sent_date = today_str
                                    print(f'Sent attendance.xlsx to {TARGET_CHAT_ID} at 21:00 WIB for {today_str}')
                                except Exception as e:
                                    print(f'Failed to send report: {e}')
                            else:
                                print('TARGET_CHAT_ID not set; skipping daily report.')
                    time.sleep(30)
                except Exception as e:
                    print(f'Scheduler error: {e}')
                    time.sleep(60)

        threading.Thread(target=send_daily_report_loop, daemon=True).start()

        # Robust polling loop: keep pending updates and retry after network issues
        while True:
            try:
                bot.infinity_polling(
                    timeout=60,
                    long_polling_timeout=60,
                    allowed_updates=ALLOWED_UPDATES,
                    skip_pending=False,  # fetch commands yang terkirim saat bot offline
                )
            except Exception as e:
                print(f'Polling error: {e}; retry in 5s')
                time.sleep(5)

    except KeyboardInterrupt:
        print("\n⚠️  Bot dihentikan oleh user")
    finally:
        if singleton:
            singleton.release()
